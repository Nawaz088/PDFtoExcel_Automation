from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import openpyxl
import os
import re
import time
import pandas as pd

def convert(pdf_file_path):
    chromedriver_path = 'C:\\Program Files\\chromedriver-win64\\chromedriver.exe'

    driver = webdriver.Chrome(executable_path=chromedriver_path)
    driver.get("https://www.pdftoexcel.com/")

    wait = WebDriverWait(driver, 30)
    file_input = wait.until(EC.presence_of_element_located((By.NAME, "Filedata")))
    file_input.send_keys(pdf_file_path)

    wait = WebDriverWait(driver, 20)
    wait.until(EC.visibility_of_element_located((By.XPATH, "//a[contains(@class, 'free-download')]")))

    download_link = driver.find_element(By.XPATH, "//a[contains(@class, 'free-download')]")
    download_link.click()

    downloaded_file_path = 'Deposit_slip.xlsx'
    #wait.until(lambda driver: not os.path.exists(downloaded_file_path + ".crdownload") and not os.path.exists(downloaded_file_path + ".tmp"))
    time.sleep(20)
    #os.rename(downloaded_file_path + ".crdownload", downloaded_file_path)
    #os.rename(downloaded_file_path + ".tmp", downloaded_file_path)

    driver.quit()

def is_valid_date(date_str):
    date_pattern = re.compile(r'\d{2}/\d{2}/\d{4}$')
    return re.match(date_pattern, date_str)

def filter_xlsx_by_date(input_file_path, output_file_path):
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active

    filtered_workbook = openpyxl.Workbook()
    filtered_sheet = filtered_workbook.active

    column_names = ["DATE", "BT", "PROV", "NAME", "BANK", "CHEQUE#", "AMOUNT", "PAYER"]
    filtered_sheet.append(column_names)

    for row in sheet.iter_rows(min_row=1, values_only=True):
        date_value = str(row[0])

        if is_valid_date(date_value):
            filtered_sheet.append(row)

    filtered_workbook.save(output_file_path)
    print(f"Filtered data saved to {output_file_path}")

def fix_leaked_amount_column(input_file_path, output_file_path):
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=7, max_col=8):
        amount_value = row[0].value
        leaked_value = row[1].value

        if not amount_value and leaked_value:
            row[0].value = leaked_value
            row[1].value = None

    workbook.save(output_file_path)
    print(f"Fixed data saved to {output_file_path}")

def extract_payment_method_names(input_file, output_file):
    Payer_Pattern = re.compile(r'[a-zA-Z\.\s]+\sPayment|[a-zA-Z\.\s]+\sPAYMENT|EFT- Patient|CHEQUE-OTHER SOURCE/ETRANSFER|COLLECTION AGENCY CHEQUE|AMERICAN EXPRESS')

    workbook = openpyxl.load_workbook(input_file)
    worksheet = workbook.active

    payment_methods_column = worksheet.max_column + 1

    previous_payment_method = None
    for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=3):
        payment_methods = []

        for cell in row:
            cell_value = cell.value
            if cell_value is not None:
                cell_value_str = str(cell_value)
                matches = Payer_Pattern.findall(cell_value_str)
                payment_methods.extend(matches)

        payment_methods_str = ', '.join(payment_methods)

        if payment_methods_str:
            previous_payment_method = payment_methods_str
        elif previous_payment_method:
            payment_methods_str = previous_payment_method

        worksheet.cell(row=row[0].row, column=payment_methods_column, value=payment_methods_str)

    workbook.save(output_file)

def shift_amount_column(file_path):
    df = pd.read_excel(file_path)

    for i in range(len(df)):
        if pd.isna(df.at[i, 'AMOUNT']):
            for j in range(6, -1, -1):  # Start from 6 (last column) and move to 0 (first column)
                if not pd.isna(df.at[i, df.columns[j]]):
                    df.at[i, 'AMOUNT'] = df.at[i, df.columns[j]]
                    df.at[i, df.columns[j]] = None
                    break  # Stop shifting once a value is found and shifted

    output_file_path = 'Step4.xlsx'
    df.to_excel(output_file_path, index=False)
    print(f"Modified data saved to {output_file_path}")

def add_summary_to_excel(file_path):
    df = pd.read_excel(file_path)

# Calculate total amount
    total_amount = df["AMOUNT"].sum()

    # Group and calculate total amount for each group
    grouped = df.groupby("Unnamed: 9")["AMOUNT"].sum()

    # Load the existing Excel file
    book = load_workbook(file_path)
    sheet_name = "Sheet1"  # Change this to your sheet name
    sheet = book[sheet_name]

    # Calculate the starting row for the summary table
    start_row = sheet.max_row + 2

    # Write the header and format
    header = ["Group Name", "Total Amount"]
    for col_num, column_title in enumerate(header, 1):
        cell = sheet.cell(row=start_row, column=col_num, value=column_title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Write the data and format
    for idx, (group, amount) in enumerate(grouped.items(), start=start_row + 1):
        sheet.cell(row=idx, column=1, value=group)
        sheet.cell(row=idx, column=2, value=amount)

    # Write the total amount and format
    sheet.cell(row=idx + 1, column=1, value="Total Amount")
    sheet.cell(row=idx + 1, column=2, value=total_amount)
    sheet.cell(row=idx + 1, column=2).font = Font(bold=True)

    # Save the changes
    book.save(file_path)
    print(f"\nOutput table added to the existing Excel {file_path}.")


def main():
    pdf_file_path = filedialog.askopenfilename(title="Select PDF File", filetypes=[("PDF files", "*.pdf")])
    convert(pdf_file_path)
    time.sleep(10)
    
    downloads_folder = os.path.expanduser('~') + '\\Downloads'
    file_name = input("give the file Name")
# Construct the full path to the downloaded Excel file
    downloaded_file_path = os.path.join(downloads_folder, file_name+".xlsx")
    print(downloaded_file_path)
    if os.path.exists(downloaded_file_path):
        input_file = downloaded_file_path
        output_file = 'Step1.xlsx'
        extract_payment_method_names(input_file, output_file)
        time.sleep(10)
        
        input_xlsx_file = 'Step1.xlsx'
        output_xlsx_file = 'Step2.xlsx'
        filter_xlsx_by_date(input_xlsx_file, output_xlsx_file)
        time.sleep(10)
        
        input_xls = 'Step2.xlsx'
        output_xls = "Step3.xlsx"
        fix_leaked_amount_column(input_xls, output_xls)

        input_fil = 'Step3.xlsx'
        shift_amount_column(input_fil)
        final_infile = 'Step4.xlsx'
        # final_outfile = 'Step5.xlsx'
        add_summary_to_excel(final_infile)
        
    else:
        print("Error: Downloaded Excel file not found.")

if __name__ == "__main__":
    main()

